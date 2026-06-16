"""
抓取维基百科「全国重点文物保护单位列表」并生成 Excel 表格
URL: https://zh.wikipedia.org/zh-cn/全国重点文物保护单位列表

使用方法:
    pip install requests beautifulsoup4 openpyxl
    python scrape_wikipedia_cultural_relics.py
"""

import re
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

BASE_URL = "https://zh.wikipedia.org/w/api.php"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/125.0.0.0 Safari/537.36"
)

HEADERS = {
    "User-Agent": USER_AGENT,
    "Accept": "application/json",
}

OUTPUT_FILE = "全国重点文物保护单位.xlsx"

COLUMNS = ["分类", "批次", "名称", "时代", "地址", "备注"]


def fetch_page_html():
    """通过 Wikipedia API 获取页面的 HTML 内容"""
    params = {
        "action": "parse",
        "page": "全国重点文物保护单位列表",
        "prop": "text",
        "format": "json",
        "formatversion": "2",
        "uselang": "zh-cn",
    }
    resp = requests.get(BASE_URL, params=params, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    data = resp.json()
    if "error" in data:
        raise RuntimeError(f"API 错误: {data['error']}")
    return data["parse"]["text"]


def clean_cell(text):
    """清洗表格单元格文本"""
    text = text.strip()
    text = re.sub(r'\[\d+\]', '', text)
    text = re.sub(r'\[注\s*\d+\]', '', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def parse_tables(html):
    """
    解析 HTML：根据 h3 章节标题确定分类，然后解析对应的 wikitable
    返回 list[dict]
    """
    soup = BeautifulSoup(html, "html.parser")
    all_rows = []

    headings = soup.find_all("div", class_="mw-heading")
    tables = soup.find_all("table", class_="wikitable")

    print(f"找到 {len(headings)} 个章节标题, {len(tables)} 个 wikitable")

    if len(tables) == 0:
        print("⚠ 未找到 wikitable，尝试宽松匹配...")
        tables = soup.find_all("table", {"class": re.compile(r".*wikitable.*")})

    heading_order = [h.get_text(strip=True).replace("[编辑]", "") for h in headings]
    category_headings = [
        h for h in heading_order
        if h not in ("分类列表", "参考文献", "外部链接")
    ]
    print(f"分类标题: {category_headings}")

    if len(category_headings) != len(tables):
        print(f"⚠ 分类数({len(category_headings)})与表格数({len(tables)})不匹配，按顺序匹配")

    for idx, table in enumerate(tables):
        category = category_headings[idx] if idx < len(category_headings) else "未知"
        is_guibing = (category == "归并名单")

        rows = table.find_all("tr")
        data_started = False

        for row in rows:
            cells = row.find_all(["th", "td"])
            if not cells:
                continue

            cell_texts = [clean_cell(cell.get_text()) for cell in cells]

            if not data_started:
                joined = "".join(cell_texts)
                if "名称" in joined and ("时代" in joined or "地址" in joined):
                    data_started = True
                continue

            if not cell_texts or all(t == "" for t in cell_texts):
                continue

            col_count = len(cell_texts)

            if is_guibing and col_count >= 5:
                name_val = cell_texts[0]
                era_val = cell_texts[1]
                addr_val = cell_texts[2]
                remark_val = cell_texts[3]
                batch_val = cell_texts[4]
            elif col_count >= 4:
                name_val = cell_texts[0]
                era_val = cell_texts[1]
                addr_val = cell_texts[2]
                batch_val = cell_texts[3]
                remark_val = ""
            else:
                continue

            if not name_val or name_val in ("—", "－", "-"):
                continue

            all_rows.append({
                "分类": category,
                "批次": batch_val,
                "名称": name_val,
                "时代": era_val,
                "地址": addr_val,
                "备注": remark_val,
            })

        print(f"  [{category}] 解析到 {len([r for r in all_rows if r['分类'] == category])} 条")

    return all_rows


def write_excel(rows, filename):
    """将数据写入 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "全国重点文物保护单位"

    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    col_widths = {
        "分类": 22, "批次": 10, "名称": 40,
        "时代": 22, "地址": 50, "备注": 40,
    }
    for col_idx, col_name in enumerate(COLUMNS, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = col_widths.get(col_name, 20)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(filename)
    print(f"\n✅ 成功保存 {len(rows)} 条记录到 {filename}")


def main():
    print("=" * 60)
    print("全国重点文物保护单位 - 维基百科数据抓取")
    print("=" * 60)

    print("\n📡 正在获取页面数据...")
    try:
        html = fetch_page_html()
        print("   ✓ 通过 API 获取成功")
    except Exception as e:
        print(f"   ✗ 获取失败: {e}")
        return

    print("\n🔍 正在解析表格数据...")
    rows = parse_tables(html)
    print(f"\n   ✓ 共解析到 {len(rows)} 条记录")

    if rows:
        from collections import Counter
        batch_counts = Counter(r["批次"] for r in rows)
        cat_counts = Counter(r["分类"] for r in rows)

        print("\n📊 各分类统计:")
        for cat, cnt in cat_counts.most_common():
            print(f"   {cat}: {cnt} 条")

        print("\n📊 各批次统计:")
        for batch, cnt in sorted(batch_counts.items()):
            print(f"   {batch}: {cnt} 条")

        write_excel(rows, OUTPUT_FILE)
    else:
        print("\n⚠ 未能解析到任何数据")
        with open("debug_output.html", "w", encoding="utf-8") as f:
            f.write(html)
        print("   HTML 已保存到 debug_output.html，请检查页面结构")


if __name__ == "__main__":
    main()